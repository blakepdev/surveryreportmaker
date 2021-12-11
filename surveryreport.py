from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

exportName = input("Date for export: ")
workbook = load_workbook(filename="...")
sheet = workbook.active

sheet.delete_rows(2)
sheet.delete_cols(4,6)
sheet.delete_cols(1)
sheet.insert_cols(1)
sheet.delete_cols(2)

workbook.save(filename=exportName + "_SurveryExport.xlsx")
